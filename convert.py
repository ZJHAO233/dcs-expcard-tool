"""
DCS Converter - Excel试验条件表转Markdown工具
"""

import re
import openpyxl


class DCSConverter:

    SKIP_HEADERS = ['序号', '条件确认']
    SECTION_HEADERS = ['试验条件', '试验恢复', '结论', '存在问题']
    LOGIC_SEPARATORS = ['与', '或', '或取反']

    def __init__(self):
        self.output = []
        self.in_content_area = False
        self.skip_section = False
        self.current_sub_title_logic = None
        self.processed_rows = set()

    def convert(self, file_path, sheet_index=0):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.worksheets[sheet_index]

        rows = []
        for r in range(1, ws.max_row + 1):
            row = []
            for c in range(1, min(ws.max_column + 1, 8)):
                val = ws.cell(r, c).value
                row.append(str(val).strip() if val else '')
            rows.append(row)

        self.output = []
        self.in_content_area = False
        self.skip_section = False
        self.current_sub_title_logic = None
        self.processed_rows = set()

        for i, row in enumerate(rows):
            if i in self.processed_rows:
                continue
            self._process_row(row, rows, i)

        return '\n'.join(self.output)

    def _process_row(self, row, all_rows, index):
        a, b, c = row[0], row[1], row[2] if len(row) > 2 else ''

        if not a and not b:
            return
        if self._is_header_row(row):
            return

        # 二级标题
        if self._is_chinese_number_title(a):
            self.output.append('')
            self.output.append('## ' + a)
            self.in_content_area = False
            self.skip_section = False
            return

        # 段落标题
        if a in self.SECTION_HEADERS:
            if a == '试验条件':
                self.skip_section = True
            self.in_content_area = False
            return

        # 进入内容区域
        if a == '试验内容':
            self.in_content_area = True
            self.skip_section = False
            return

        if self.skip_section:
            return

        # 三级标题
        if self._is_sub_title(a):
            # 统一括号格式为全角
            normalized = a.replace('(', '（').replace(')', '）')
            self.output.append('')
            self.output.append('### ' + normalized)
            self.output.append('')
            self.current_sub_title_logic = self._extract_title_logic(normalized)
            return

        # 子列一级
        if a.isdigit() and '.' not in a:
            self._process_level1(a, b, c, row, all_rows, index)
            return

        # 子列二级
        if self._is_sub_item(a):
            self._process_level2(a, b, c, row, all_rows, index)
            return

    def _is_header_row(self, row):
        a, b = row[0], row[1]
        if a in self.SKIP_HEADERS:
            return True
        if a == '序号' and ('条件' in b or '内容' in b):
            return True
        return False

    def _is_chinese_number_title(self, text):
        return bool(re.match(r'^[一二三四五六七八九十]+、', text))

    def _is_sub_title(self, text):
        return bool(re.match(r'^\d+\.[\u4e00-\u9fa5]', text))

    def _is_sub_item(self, text):
        return bool(re.match(r'^\d+\.\d+$', text))

    def _extract_title_logic(self, title):
        match = re.search(r'[（(]([与或])[）)]', title)
        return match.group(1) if match else None

    def _is_logic_separator(self, text):
        return text in self.LOGIC_SEPARATORS

    def _get_content_from_row(self, row, start_col):
        for i in range(start_col, len(row)):
            val = row[i]
            if val and not self._is_logic_separator(val):
                return val
        return None

    def _process_level1(self, a, b, c, row, all_rows, index):
        # 判断是否有子项
        has_children = False
        child_logic = None
        if index + 1 < len(all_rows):
            next_a = all_rows[index + 1][0]
            if self._is_sub_item(next_a) and next_a.startswith(a + '.'):
                has_children = True
                # 从第一个子项的B列获取逻辑
                child_logic = all_rows[index + 1][1] if self._is_logic_separator(all_rows[index + 1][1]) else None

        # 获取内容
        content = self._get_content_from_row(row, start_col=1)
        if not content:
            content = b if b and not self._is_logic_separator(b) else ''

        # 判断是否输出逻辑
        logic_str = ''
        if has_children and child_logic:
            # 只有当子列一级自己的B列逻辑与三级标题逻辑一致时才忽略
            # 如果子列一级B列不是逻辑分隔符，则使用子项的逻辑
            if self._is_logic_separator(b):
                # 子列一级B列是逻辑分隔符，检查是否与三级标题一致
                if b != self.current_sub_title_logic:
                    logic_str = '（' + b + '）'
            else:
                # 子列一级B列不是逻辑分隔符，使用子项的逻辑
                logic_str = '（' + child_logic + '）'
        
        # 处理或取反
        if b and '或取反' in b:
            logic_str = '（或取反）'

        self.output.append(a + '. ' + content + logic_str)

    def _process_level2(self, a, b, c, row, all_rows, index):
        # C列是子列三级逻辑（标记新分组，也是延续行使用的逻辑）
        child_logic = c if self._is_logic_separator(c) else None

        # 获取内容
        # 如果C列是逻辑分隔符，从D列开始找内容
        # 如果C列不是逻辑分隔符，从B列开始找内容
        content = None
        if child_logic:
            # C列是逻辑分隔符，从D列开始找内容
            for col_idx in range(3, len(row)):
                val = row[col_idx]
                if val and not self._is_logic_separator(val):
                    content = val
                    break
        else:
            # C列不是逻辑分隔符，从B列开始找内容
            for col_idx in range(1, len(row)):
                val = row[col_idx]
                if val and not self._is_logic_separator(val):
                    content = val
                    break

        if not content:
            return

        if child_logic:
            # 有逻辑分隔符，开始新分组
            group_items = [content]
            # 跟踪延续行使用的逻辑（使用C列的逻辑）
            continuation_logic = child_logic
            j = index + 1
            while j < len(all_rows):
                next_row = all_rows[j]
                next_a = next_row[0]
                
                # 检查是否是同级子项（有编号的）
                if self._is_sub_item(next_a) and next_a.startswith(a.split('.')[0] + '.'):
                    next_c = next_row[2] if len(next_row) > 2 else ''
                    if self._is_logic_separator(next_c):
                        break
                    # 获取内容
                    next_content = None
                    for col_idx in range(2, len(next_row)):
                        val = next_row[col_idx]
                        if val and not self._is_logic_separator(val):
                            next_content = val
                            break
                    if next_content:
                        group_items.append(next_content)
                        self.processed_rows.add(j)
                    j += 1
                # 检查是否是无编号的延续行（A列为空，C列有内容）
                elif not next_a and len(next_row) > 2:
                    next_c = next_row[2] if len(next_row) > 2 else ''
                    if self._is_logic_separator(next_c):
                        break
                    # 获取内容
                    next_content = None
                    for col_idx in range(2, len(next_row)):
                        val = next_row[col_idx]
                        if val and not self._is_logic_separator(val):
                            next_content = val
                            break
                    if next_content:
                        group_items.append(next_content)
                        self.processed_rows.add(j)
                    j += 1
                else:
                    break

            if len(group_items) > 1:
                result = group_items[0]
                # 根据C列的逻辑使用对应的分隔符
                separator = ' 且 ' if continuation_logic == '与' else ' 或 '
                for item in group_items[1:]:
                    result = result + separator + item
                self.output.append('   - ' + result)
            else:
                self.output.append('   - ' + content)
        else:
            self.output.append('   - ' + content)


def convert_sheet(file_path, sheet_index=0):
    converter = DCSConverter()
    return converter.convert(file_path, sheet_index)


if __name__ == '__main__':
    import sys

    if len(sys.argv) < 2:
        print("Usage: python convert.py <excel_file> [sheet_index]")
        sys.exit(1)

    file_path = sys.argv[1]
    sheet_index = int(sys.argv[2]) if len(sys.argv) > 2 else 0

    try:
        result = convert_sheet(file_path, sheet_index)
        with open('output.md', 'w', encoding='utf-8') as f:
            f.write(result)
        print("Done! Output written to output.md")
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
